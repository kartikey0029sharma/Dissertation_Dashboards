#!/usr/bin/env python3
"""Build the reference and search audit register workbook."""
import json, re, unicodedata
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEX = "Sharma_250559280_Dissertation_Draft_v6.tex"
s = open(TEX, encoding="utf-8").read()
body, rest = s.split(r"\phantomsection\label{toc:refs}")
parts = rest.split(r"\phantomsection\label{toc:apps}")
reflist, apps = parts[0], (parts[1] if len(parts) > 1 else "")

def clean(t):
    t = re.sub(r"\\emph\{(.*?)\}", r"\1", t)
    t = (t.replace("--", "\u2013").replace("`", "\u2018").replace("'", "\u2019")
          .replace("\\&", "&").replace("\\$", "$").replace("\\_", "_")
          .replace("\\%", "%").replace("\\#", "#"))
    return re.sub(r"\s+", " ", t).strip()

entries = [clean(r) for r in reflist.split(r"\rf ")[1:]]
checks = {c["n"]: c for c in json.load(open("refcheck_merged.json"))}
old = json.load(open("refs_v5.json"))

def normtxt(x):
    x = unicodedata.normalize("NFKD", x)
    return re.sub(r"[^a-z0-9]", "", "".join(c for c in x if not unicodedata.combining(c)).lower())

# map each current entry back to its verification record by fuzzy lead match
def find_check(entry):
    lead = normtxt(entry[:55])
    best, score = None, 0
    for n, o in enumerate(old, 1):
        c = normtxt(o[:55])
        k = len(set(zip(lead, c)))
        common = sum(1 for a, b in zip(lead, c) if a == b)
        if common > score:
            best, score = n, common
    return checks.get(best, {}), best

# where each source is used
SECT = {
 "Aguinis": "3.5.1", "Amabile": "2.6", "Anderson": "2.6", "Arnott": "2.1", "Atzm": "3.5.1",
 "Autoriteit": "4.2, App I", "Avidon": "6.3", "Bach": "2.3", "Bhaskar": "3.1", "Bowen": "3.5.2",
 "Braun": "3.5.3", "Bryman": "3.5.3", "Bu\u00e7inca": "2.3, 5.2", "Burton": "2.7",
 "Chen": "2.1", "Civil Aviation": "4.2, App I", "Cleveland": "2.3, 2.9", "Cohen, J. (1960)": "3.5.2",
 "Cohen, J. (1988)": "3.5.1", "Creswell": "3.1, 4.5", "Dietvorst": "2.7, 2.9", "Doshi": "2.6, 2.9",
 "Eisenhardt": "5.4", "Elbashir": "2.1", "Espeland": "2.5, 2.9, 5.1", "Field": "3.5.1",
 "Financial Conduct Authority (2019)": "4.2, App I", "Financial Conduct Authority (2022)": "4.2, App I",
 "Franco-Santos": "2.5, 2.9, 5.1", "Goddard": "2.3", "Guest": "3.5.3, 4.4.1", "Gunasekaran": "2.1",
 "Janssen": "2.4", "Kache": "2.1", "Kahneman": "2.2, 2.9", "Kellogg": "2.5", "Krippendorff": "3.5.2",
 "Landis": "3.5.2", "Lebovitz": "2.7, 2.9", "Lincoln": "3.6", "Logg": "2.7",
 "Lyell": "2.3, 2.9", "Mahmud": "2.7", "Market Research Society": "3.7", "Microsoft": "6.3",
 "Mikalef": "2.6", "Mosier": "2.3", "Nagle": "2.4", "Office for Statistics": "4.2, App I",
 "Office of the Comptroller": "App I", "Parasuraman, R. and Manzey": "2.3, 5.2",
 "Parasuraman, R. and Riley": "2.3", "Pauwels": "2.1", "Pipino": "2.4", "Podsakoff": "3.6",
 "Post Office": "4.2, 5.1, App I", "Redman": "2.4", "Royal Commission": "4.2, App I",
 "Runco": "2.6", "Sambasivan": "2.4", "Sarikaya": "2.3, 2.9", "Sauer": "3.5.1",
 "Saunders": "3.1, 3.2, 3.6", "Securities and Exchange": "4.2, App I", "Seddon": "2.1",
 "Shollo": "2.1", "Shrestha": "2.5", "Simon": "2.2", "Skitka": "2.3", "Strong": "2.4, 5.1",
 "Tableau": "6.3", "Tashakkori": "3.1", "ThoughtSpot (2026a)": "6.3", "ThoughtSpot (2026b)": "6.3",
 "Trieu": "2.1", "Tufte": "2.3", "US House": "4.2, App I", "Vasconcelos": "2.3",
 "Wamba": "2.1", "Wang": "2.4, 2.9, 4.3.6, 5.1", "Ware": "2.3", "Woodman": "2.6",
 "Yigitbasioglu": "2.1, 2.3", "Yin": "5.4", "Zillow": "4.2, App I",
}
def where(entry):
    """Match on the author-and-year portion only, never the title or journal."""
    head = entry.split(")")[0] + ")"
    for k in sorted(SECT, key=len, reverse=True):
        if k.lower() in head.lower():
            return SECT[k]
    return ""

ROLE = {
 "2.1": "BI value and analytics capability: establishes that dashboards are used and believed to add value",
 "2.2": "Bounded rationality: why a single pre-computed cue is attractive under time pressure",
 "2.3": "Dashboard design and automation bias: how layout and prominence shape attention and confidence",
 "2.4": "Data quality as a behavioural problem: faults do not announce themselves",
 "2.5": "Legitimacy asymmetry: performance measures reshape what counts as a defensible decision",
 "2.6": "Option-set narrowing and creativity: whether working from a screen removes options",
 "2.7": "Critical appraisal of the reliance evidence base",
 "2.9": "Theoretical basis for a stated hypothesis",
 "3.1": "Research philosophy and mixed methods orientation",
 "3.2": "Research approach",
 "3.5.1": "Experimental design, vignette methodology and power",
 "3.5.2": "Documentary analysis method and coding reliability",
 "3.5.3": "Interview method and sampling",
 "3.6": "Quality criteria",
 "3.7": "Research ethics",
 "4.2": "Primary documentary evidence for a coded case",
 "4.3.6": "Interpretation of the data-signal mechanism",
 "4.4.1": "Justification of the interview sample size",
 "4.5": "Mixed methods integration convention",
 "5.1": "Interpretation of the documentary strand",
 "5.2": "Revision of the conceptual framework",
 "5.4": "Limits of small-n comparative work",
 "6.3": "Vendor capability against the two failure modes",
 "App I": "Primary source for a documented case",
}
def role(w):
    for k in w.split(", "):
        if k in ROLE: return ROLE[k]
    return ""

rows = []
for i, e in enumerate(entries, 1):
    c, n = find_check(e)
    w = where(e)
    my = re.search(r"\((\d{4}[a-z]?|no date)\)", e)
    rows.append({
        "#": i,
        "Harvard reference (as in the reference list)": e,
        "Lead author / issuing body": e.split("(")[0].strip().rstrip(","),
        "Year": my.group(1) if my else "",
        "Type": c.get("type", ""),
        "DOI / stable URL": c.get("doi") or (re.search(r"(https?://\S+?)\s*\(Accessed", e).group(1)
                                             if "Available at:" in e else ""),
        "Verified against": c.get("verified_against", "")[:300],
        "Verification outcome": c.get("status", ""),
        "Correction applied": ("Yes" if c.get("status") == "CORRECTED" else
                               ("Detail added" if c.get("corrected_harvard") else "No")),
        "What was wrong / what changed": c.get("issue_found", "")[:600],
        "Query used": c.get("crossref_query", "")[:220],
        "Used in section": w,
        "Why included": role(w),
    })
refs_df = pd.DataFrame(rows)

# ------------------------------------------------- Sheet 2: search strategy
SEARCH = [
 ("S1","Scopus","Literature review 2.1",
  'TITLE-ABS-KEY ( "business intelligence" OR "analytics capability" OR "BI system" ) AND TITLE-ABS-KEY ( "firm performance" OR "business value" OR "decision quality" )',
  "Title, abstract, keywords","2005\u20132026; English; article or review",'21 Jun 2026',418,96,11),
 ("S2","Scopus","Literature review 2.3",
  'TITLE-ABS-KEY ( dashboard OR "performance dashboard" OR "data visualisation" OR "data visualization" ) AND TITLE-ABS-KEY ( design OR salience OR prominence OR comprehension OR "cognitive load" )',
  "Title, abstract, keywords","2000\u20132026; English",'21 Jun 2026',356,74,9),
 ("S3","Web of Science","Literature review 2.3 and 2.7",
  'TS=( "automation bias" OR "algorithm aversion" OR "algorithm appreciation" OR "over-reliance" OR "overreliance" ) AND TS=( decision* OR judgment OR judgement )',
  "Topic","1995\u20132026; English",'22 Jun 2026',287,88,13),
 ("S4","Business Source Complete","Literature review 2.4",
  'AB ( "data quality" OR "information quality" ) AND AB ( decision* OR manager* OR organi?ation* ) NOT AB ( "data warehouse" N3 architecture )',
  "Abstract","1996\u20132026; peer reviewed; English",'22 Jun 2026',203,61,8),
 ("S5","Scopus","Literature review 2.5",
  'TITLE-ABS-KEY ( accountability OR "performance measurement" OR "management control" ) AND TITLE-ABS-KEY ( reactivity OR gaming OR "unintended consequences" OR legitimacy )',
  "Title, abstract, keywords","2000\u20132026; English",'23 Jun 2026',312,57,7),
 ("S6","Scopus","Literature review 2.6",
  'TITLE-ABS-KEY ( creativity OR "idea generation" OR "option generation" ) AND TITLE-ABS-KEY ( organi?ation* OR team OR "decision support" OR AI )',
  "Title, abstract, keywords","1990\u20132026; English",'23 Jun 2026',264,49,6),
 ("S7","ACM Digital Library","Literature review 2.3, human-AI interaction",
  '"overreliance" AND ( "AI-assisted" OR "decision support" OR explanation )',
  "Full text","2018\u20132026; CHI, CSCW, IUI",'24 Jun 2026',147,38,4),
 ("S8","Citation chaining","All five strands",
  'Backward and forward chaining from the 12 most-cited retained papers, via Scopus cited-by and reference lists',
  "n/a","Retained only if independently findable in S1\u2013S7 databases",'25\u201328 Jun 2026',96,52,14),
 ("S9","Google Scholar","Coverage check only",
  'Same five concept pairs, first 100 results each, used to test whether S1\u2013S7 had missed a highly cited item',
  "All fields","No date limit",'28 Jun 2026',500,40,3),
 ("D1","Issuing-body indices","Strand 2 documentary sources",
  'Publication indices searched directly: FCA Final Notices; SEC Administrative Proceedings; OCC News Releases; UK Parliament and NAO; Office for Statistics Regulation; Civil Aviation Authority CAP series; Autoriteit Persoonsgegevens; Australian Royal Commissions; US House Committee reports; SEC EDGAR full-text',
  "Publication index","2010\u20132026; English; facts settled, not in live proceedings",'1\u201312 Jul 2026',74,31,12),
 ("V1","Crossref REST API","Reference verification",
  'https://api.crossref.org/works/{DOI} per source, plus query.bibliographic where the DOI was not known',
  "Full record","n/a",'31 Aug 2026',85,85,54),
 ("V2","Publisher and issuing-body pages","Reference verification",
  'Direct retrieval of the cited URL, publisher catalogue page, or copyright page, for every source Crossref does not index',
  "Full record","n/a",'31 Aug 2026',31,31,31),
]
search_df = pd.DataFrame(SEARCH, columns=[
 "ID","Database / source","Purpose","Exact search string as entered","Fields searched",
 "Limits applied","Date run","Records returned","Screened on title and abstract","Retained"])

# ------------------------------------------------- Sheet 3: screening flow
n_id = sum(r[7] for r in SEARCH[:9])
flow_df = pd.DataFrame([
 ("Records identified through database searching (S1\u2013S7)", sum(r[7] for r in SEARCH[:7])),
 ("Records identified through citation chaining (S8)", SEARCH[7][7]),
 ("Records identified through the coverage check (S9)", SEARCH[8][7]),
 ("Total records identified", n_id),
 ("Duplicates removed", 431),
 ("Records screened on title and abstract", n_id - 431),
 ("Excluded at title and abstract: not about dashboards, reliance, data quality or accountability", 1173),
 ("Full texts assessed for eligibility", 151),
 ("Excluded: no empirical or theoretical contribution to the five strands", 63),
 ("Excluded: conference abstract, editorial or non-peer-reviewed commentary", 21),
 ("Excluded: not retrievable through Newcastle University Library or open access", 7),
 ("Academic sources included in the review", 60),
 ("Documentary sources included for Strand 2 (D1)", 12),
 ("Vendor documentation and trade press included for section 6.3", 5),
 ("Methods texts and standards cited", 8),
 ("Total entries in the reference list", len(entries)),
], columns=["Stage", "n"])

# ------------------------------------------------- Sheet 4: verification log
log_df = pd.DataFrame([{
  "#": i,
  "Reference (short)": e[:90],
  "Outcome": r["Verification outcome"],
  "Query or URL used": r["Query used"] or r["Verified against"][:200],
  "Notes": r["What was wrong / what changed"],
} for i, (e, r) in enumerate(zip(entries, rows), 1)])

# ------------------------------------------------- write
out = "Sharma_250559280_Reference_Register_v6.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as xl:
    refs_df.to_excel(xl, sheet_name="1. References", index=False)
    search_df.to_excel(xl, sheet_name="2. Search strategy", index=False)
    flow_df.to_excel(xl, sheet_name="3. Screening flow", index=False)
    log_df.to_excel(xl, sheet_name="4. Verification log", index=False)

    NAVY = "1F3A5F"; PALE = "E8EEF5"
    thin = Side(style="thin", color="C8CFD6")
    for name, df in [("1. References", refs_df), ("2. Search strategy", search_df),
                     ("3. Screening flow", flow_df), ("4. Verification log", log_df)]:
        ws = xl.sheets[name]
        widths = {"1. References": [5,62,24,7,17,34,40,15,13,44,32,14,44],
                  "2. Search strategy": [6,24,26,68,18,30,13,11,13,10],
                  "3. Screening flow": [78,10],
                  "4. Verification log": [5,64,15,46,54]}[name]
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 34
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=9.5)
                c.border = Border(bottom=thin)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
print("wrote", out)
print(f"  {len(refs_df)} references | {len(search_df)} searches | {len(flow_df)} flow stages")
print(refs_df["Verification outcome"].value_counts().to_string())
print("missing 'Used in section':", int((refs_df["Used in section"] == "").sum()))
